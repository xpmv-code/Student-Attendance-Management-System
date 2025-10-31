from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required
from app import db
from app.models import Student
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
from io import BytesIO
from datetime import datetime

# 创建学生管理蓝图
student_bp = Blueprint('student', __name__)

# 允许的文件扩展名
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@student_bp.route('/')
@login_required
def index():
    """学生列表页面"""
    # 获取查询参数
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    political_status = request.args.get('political_status', '', type=str)
    
    # 构建查询
    query = Student.query
    
    # 搜索条件
    if search:
        query = query.filter(
            (Student.student_id.contains(search)) |
            (Student.student_name.contains(search)) |
            (Student.phone.contains(search))
        )
    
    # 政治面貌筛选
    if political_status:
        query = query.filter(Student.political_status == political_status)
    
    # 分页查询
    pagination = query.order_by(Student.create_time.desc()).paginate(
        page=page,
        per_page=10,
        error_out=False
    )
    
    students = pagination.items
    
    # 获取所有政治面貌选项（用于筛选下拉框）
    political_statuses = db.session.query(Student.political_status)\
        .distinct()\
        .filter(Student.political_status.isnot(None))\
        .order_by(Student.political_status)\
        .all()
    political_statuses = [ps[0] for ps in political_statuses]
    
    return render_template(
        'student/index.html',
        students=students,
        pagination=pagination,
        search=search,
        political_status=political_status,
        political_statuses=political_statuses
    )


@student_bp.route('/add', methods=['GET', 'POST'])
@login_required
def add():
    """添加学生"""
    if request.method == 'POST':
        student_id = request.form.get('student_id', '').strip()
        student_name = request.form.get('student_name', '').strip()
        political_status = request.form.get('political_status', '').strip()
        phone = request.form.get('phone', '').strip()
        
        # 数据验证
        if not student_id:
            flash('学号不能为空！', 'error')
            return redirect(url_for('student.add'))
        
        if not student_name:
            flash('姓名不能为空！', 'error')
            return redirect(url_for('student.add'))
        
        if not phone:
            flash('联系电话不能为空！', 'error')
            return redirect(url_for('student.add'))
        
        # 检查学号是否已存在
        existing_student = Student.query.filter_by(student_id=student_id).first()
        if existing_student:
            flash(f'学号 {student_id} 已存在！', 'error')
            return redirect(url_for('student.add'))
        
        # 创建新学生
        student = Student(
            student_id=student_id,
            student_name=student_name,
            political_status=political_status if political_status else None,
            phone=phone
        )
        
        try:
            db.session.add(student)
            db.session.commit()
            flash(f'学生 {student_name} 添加成功！', 'success')
            return redirect(url_for('student.index'))
        except Exception as e:
            db.session.rollback()
            flash(f'添加失败：{str(e)}', 'error')
            return redirect(url_for('student.add'))
    
    return render_template('student/form.html', student=None, action='add')


@student_bp.route('/edit/<student_id>', methods=['GET', 'POST'])
@login_required
def edit(student_id):
    """编辑学生信息"""
    student = Student.query.get_or_404(student_id)
    
    if request.method == 'POST':
        student_name = request.form.get('student_name', '').strip()
        political_status = request.form.get('political_status', '').strip()
        phone = request.form.get('phone', '').strip()
        
        # 数据验证
        if not student_name:
            flash('姓名不能为空！', 'error')
            return redirect(url_for('student.edit', student_id=student_id))
        
        if not phone:
            flash('联系电话不能为空！', 'error')
            return redirect(url_for('student.edit', student_id=student_id))
        
        # 更新学生信息
        student.student_name = student_name
        student.political_status = political_status if political_status else None
        student.phone = phone
        
        try:
            db.session.commit()
            flash(f'学生 {student_name} 信息更新成功！', 'success')
            return redirect(url_for('student.index'))
        except Exception as e:
            db.session.rollback()
            flash(f'更新失败：{str(e)}', 'error')
            return redirect(url_for('student.edit', student_id=student_id))
    
    return render_template('student/form.html', student=student, action='edit')


@student_bp.route('/delete/<student_id>', methods=['POST'])
@login_required
def delete(student_id):
    """删除学生"""
    student = Student.query.get_or_404(student_id)
    student_name = student.student_name
    
    try:
        db.session.delete(student)
        db.session.commit()
        flash(f'学生 {student_name} 删除成功！', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败：{str(e)}', 'error')
    
    return redirect(url_for('student.index'))


@student_bp.route('/import', methods=['GET', 'POST'])
@login_required
def import_excel():
    """从Excel导入学生信息"""
    if request.method == 'POST':
        # 检查是否有文件上传
        if 'file' not in request.files:
            flash('没有选择文件！', 'error')
            return redirect(url_for('student.import_excel'))
        
        file = request.files['file']
        
        # 检查文件名是否为空
        if file.filename == '':
            flash('没有选择文件！', 'error')
            return redirect(url_for('student.import_excel'))
        
        # 检查文件格式
        if not allowed_file(file.filename):
            flash('仅支持 .xlsx 或 .xls 格式的Excel文件！', 'error')
            return redirect(url_for('student.import_excel'))
        
        try:
            # 读取Excel文件
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            success_count = 0
            error_count = 0
            error_messages = []
            
            # 从第2行开始读取（第1行是表头）
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # 跳过空行
                if not any(row):
                    continue
                
                # 确保至少有4列数据
                if len(row) < 4:
                    error_messages.append(f'第 {row_num} 行：数据列数不足')
                    error_count += 1
                    continue
                
                student_id = str(row[0]).strip() if row[0] else ''
                student_name = str(row[1]).strip() if row[1] else ''
                political_status = str(row[2]).strip() if row[2] else ''
                phone = str(row[3]).strip() if row[3] else ''
                
                # 数据验证
                if not student_id:
                    error_messages.append(f'第 {row_num} 行：学号为空')
                    error_count += 1
                    continue
                
                if not student_name:
                    error_messages.append(f'第 {row_num} 行：姓名为空')
                    error_count += 1
                    continue
                
                if not phone:
                    error_messages.append(f'第 {row_num} 行：联系电话为空')
                    error_count += 1
                    continue
                
                # 检查学号是否已存在
                existing_student = Student.query.filter_by(student_id=student_id).first()
                if existing_student:
                    # 更新现有学生信息
                    existing_student.student_name = student_name
                    existing_student.political_status = political_status if political_status else None
                    existing_student.phone = phone
                    success_count += 1
                else:
                    # 创建新学生
                    student = Student(
                        student_id=student_id,
                        student_name=student_name,
                        political_status=political_status if political_status else None,
                        phone=phone
                    )
                    db.session.add(student)
                    success_count += 1
            
            # 提交数据库更改
            db.session.commit()
            
            # 显示导入结果
            if error_count > 0:
                flash(f'导入完成！成功 {success_count} 条，失败 {error_count} 条', 'warning')
                for msg in error_messages[:5]:  # 只显示前5条错误
                    flash(msg, 'error')
                if len(error_messages) > 5:
                    flash(f'...还有 {len(error_messages) - 5} 条错误未显示', 'error')
            else:
                flash(f'导入成功！共导入 {success_count} 条学生记录', 'success')
            
            return redirect(url_for('student.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'导入失败：{str(e)}', 'error')
            return redirect(url_for('student.import_excel'))
    
    return render_template('student/import.html')


@student_bp.route('/export')
@login_required
def export():
    """导出学生名单为Excel文件"""
    # 获取查询参数（与列表页面相同）
    search = request.args.get('search', '', type=str)
    political_status = request.args.get('political_status', '', type=str)
    
    # 构建查询（与列表页面相同的筛选逻辑）
    query = Student.query
    
    # 搜索条件
    if search:
        query = query.filter(
            (Student.student_id.contains(search)) |
            (Student.student_name.contains(search)) |
            (Student.phone.contains(search))
        )
    
    # 政治面貌筛选
    if political_status:
        query = query.filter(Student.political_status == political_status)
    
    # 获取所有符合条件的学生（不分页）
    students = query.order_by(Student.student_id).all()
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生名单"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入表头
    headers = ["学号", "姓名", "政治面貌", "联系电话", "创建时间"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15  # 学号
    ws.column_dimensions['B'].width = 12  # 姓名
    ws.column_dimensions['C'].width = 12  # 政治面貌
    ws.column_dimensions['D'].width = 15  # 联系电话
    ws.column_dimensions['E'].width = 20  # 创建时间
    
    # 写入数据
    data_alignment = Alignment(horizontal="left", vertical="center")
    for row_num, student in enumerate(students, 2):
        ws.cell(row=row_num, column=1).value = student.student_id
        ws.cell(row=row_num, column=2).value = student.student_name
        ws.cell(row=row_num, column=3).value = student.political_status or '-'
        ws.cell(row=row_num, column=4).value = student.phone
        ws.cell(row=row_num, column=5).value = student.create_time.strftime('%Y-%m-%d %H:%M:%S') if student.create_time else '-'
        
        # 设置数据行对齐
        for col_num in range(1, 6):
            ws.cell(row=row_num, column=col_num).alignment = data_alignment
    
    # 生成文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'学生名单_{timestamp}.xlsx'
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 返回文件
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@student_bp.route('/detail/<student_id>')
@login_required
def detail(student_id):
    """查看学生详情"""
    from app.models import Attendance, LeaveRecord
    
    student = Student.query.get_or_404(student_id)
    
    # 获取考勤统计
    attendance_stats = {
        'total': student.attendances.count(),
        'normal': student.attendances.filter_by(attendance_type='正常').count(),
        'late': student.attendances.filter_by(attendance_type='迟到').count(),
        'absent': student.attendances.filter_by(attendance_type='缺席').count(),
    }
    
    # 获取请假统计
    leave_stats = {
        'total': student.leave_records.count(),
    }
    
    # 获取最近的考勤记录（最多5条）
    recent_attendances = Attendance.query.filter_by(student_id=student_id)\
        .order_by(Attendance.attendance_date.desc())\
        .limit(5)\
        .all()
    
    # 获取最近的请假记录（最多5条）
    recent_leaves = LeaveRecord.query.filter_by(student_id=student_id)\
        .order_by(LeaveRecord.create_time.desc())\
        .limit(5)\
        .all()
    
    return render_template(
        'student/detail.html',
        student=student,
        attendance_stats=attendance_stats,
        leave_stats=leave_stats,
        recent_attendances=recent_attendances,
        recent_leaves=recent_leaves
    )

