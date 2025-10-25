"""
考勤记录路由模块
"""
from flask import Blueprint, render_template, request, flash, redirect, url_for, send_file
from datetime import datetime, timedelta, date
from sqlalchemy import func, and_, case

from app import db
from app.models.attendance import Attendance
from app.models.student import Student
from app.models.course import Course
from app.utils.week_helper import get_week_number

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

attendance_bp = Blueprint('attendance', __name__)

@attendance_bp.route('/')
def index():
    """考勤日历视图"""
    # 获取当前日期或请求的日期
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    try:
        current_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        current_date = date.today()
    
    # 获取当天的所有课程
    # 根据星期几查询课程
    weekday_names = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    current_weekday = weekday_names[current_date.weekday()]
    all_courses = Course.query.filter(
        Course.course_time.like(f'{current_weekday}%')
    ).order_by(Course.course_time).all()
    
    # 根据周次过滤课程
    current_week = get_week_number(current_date)
    courses = []
    if current_week:
        for course in all_courses:
            if course.is_teaching_week(current_week):
                courses.append(course)
    else:
        # 如果不在学期范围内，显示所有课程
        courses = all_courses
    
    # 为每个课程获取考勤统计
    course_stats = []
    for course in courses:
        # 获取该课程该日期的考勤记录
        attendance_count = Attendance.query.filter_by(
            course_id=course.course_id,
            attendance_date=current_date
        ).count()
        
        # 统计各种状态的人数
        present = Attendance.query.filter_by(
            course_id=course.course_id,
            attendance_date=current_date,
            attendance_type='到课'
        ).count()
        
        leave = Attendance.query.filter_by(
            course_id=course.course_id,
            attendance_date=current_date,
            attendance_type='请假'
        ).count()
        
        absent = Attendance.query.filter_by(
            course_id=course.course_id,
            attendance_date=current_date,
            attendance_type='旷课'
        ).count()
        
        # 计算到勤率
        total_students = Student.query.count()
        attendance_rate = round((present / total_students * 100), 1) if total_students > 0 else 0
        
        course_stats.append({
            'course': course,
            'total': total_students,
            'recorded': attendance_count,
            'present': present,
            'leave': leave,
            'absent': absent,
            'rate': attendance_rate,
            'is_recorded': attendance_count > 0
        })
    
    # 获取周一到周日的日期
    week_start = current_date - timedelta(days=current_date.weekday())
    week_dates = [week_start + timedelta(days=i) for i in range(7)]
    
    return render_template(
        'attendance/index.html',
        current_date=current_date,
        courses=course_stats,
        week_dates=week_dates
    )

@attendance_bp.route('/record/<course_id>')
def record(course_id):
    """考勤记录编辑页面"""
    # 获取日期参数
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    try:
        attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('无效的日期格式', 'error')
        return redirect(url_for('attendance.index'))
    
    # 获取课程信息
    course = Course.query.get_or_404(course_id)
    
    # 获取所有学生
    students = Student.query.order_by(Student.student_id).all()
    
    # 获取该课程该日期的考勤记录
    attendance_records = {}
    records = Attendance.query.filter_by(
        course_id=course_id,
        attendance_date=attendance_date
    ).all()
    
    for record in records:
        attendance_records[record.student_id] = {
            'id': record.attendance_id,
            'type': record.attendance_type,
            'late_minutes': record.late_minutes,
            'note': record.attendance_note
        }
    
    # 如果没有考勤记录，自动创建默认记录（所有学生到课）
    if not records:
        for student in students:
            new_record = Attendance(
                student_id=student.student_id,
                course_id=course_id,
                attendance_date=attendance_date,
                attendance_type='到课',
                late_minutes=0,
                attendance_note=''
            )
            db.session.add(new_record)
        
        try:
            db.session.commit()
            flash('已为该课程创建默认考勤记录（所有学生到课）', 'success')
            return redirect(url_for('attendance.record', course_id=course_id, date=date_str))
        except Exception as e:
            db.session.rollback()
            flash(f'创建默认考勤记录失败：{str(e)}', 'error')
    
    # 构建学生考勤数据
    student_records = []
    for student in students:
        if student.student_id in attendance_records:
            record = attendance_records[student.student_id]
            student_records.append({
                'student': student,
                'attendance_id': record['id'],
                'attendance_type': record['type'],
                'late_minutes': record['late_minutes'],
                'note': record['note']
            })
        else:
            student_records.append({
                'student': student,
                'attendance_id': None,
                'attendance_type': '到课',
                'late_minutes': 0,
                'note': ''
            })
    
    return render_template(
        'attendance/record.html',
        course=course,
        attendance_date=attendance_date,
        student_records=student_records
    )

@attendance_bp.route('/save', methods=['POST'])
def save():
    """保存考勤记录"""
    course_id = request.form.get('course_id')
    date_str = request.form.get('attendance_date')
    
    try:
        attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('无效的日期格式', 'error')
        return redirect(url_for('attendance.index'))
    
    # 获取所有学生
    students = Student.query.all()
    
    # 更新考勤记录
    for student in students:
        attendance_type = request.form.get(f'attendance_type_{student.student_id}', '到课')
        late_minutes = request.form.get(f'late_minutes_{student.student_id}', 0)
        note = request.form.get(f'note_{student.student_id}', '')
        
        # 处理迟到时间
        try:
            late_minutes = int(late_minutes) if late_minutes else 0
        except ValueError:
            late_minutes = 0
        
        # 查找现有记录
        record = Attendance.query.filter_by(
            student_id=student.student_id,
            course_id=course_id,
            attendance_date=attendance_date
        ).first()
        
        if record:
            # 更新现有记录
            record.attendance_type = attendance_type
            record.late_minutes = late_minutes
            record.attendance_note = note
        else:
            # 创建新记录
            new_record = Attendance(
                student_id=student.student_id,
                course_id=course_id,
                attendance_date=attendance_date,
                attendance_type=attendance_type,
                late_minutes=late_minutes,
                attendance_note=note
            )
            db.session.add(new_record)
    
    try:
        db.session.commit()
        flash('考勤记录保存成功', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'保存失败：{str(e)}', 'error')
    
    return redirect(url_for('attendance.record', course_id=course_id, date=date_str))

@attendance_bp.route('/export_day')
def export_day():
    """导出当天的考勤记录"""
    # 获取日期参数
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    try:
        export_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('无效的日期格式', 'error')
        return redirect(url_for('attendance.index'))
    
    # 获取当天的所有课程
    # 根据星期几查询课程
    weekday_names = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    export_weekday = weekday_names[export_date.weekday()]
    all_courses = Course.query.filter(
        Course.course_time.like(f'{export_weekday}%')
    ).order_by(Course.course_time).all()
    
    # 根据周次过滤课程
    export_week = get_week_number(export_date)
    courses = []
    if export_week:
        for course in all_courses:
            if course.is_teaching_week(export_week):
                courses.append(course)
    else:
        courses = all_courses
    
    if not courses:
        flash('该日期没有课程', 'warning')
        return redirect(url_for('attendance.index'))
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = f"{export_date.strftime('%Y-%m-%d')}考勤"
    
    # 定义样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入标题
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{export_date.strftime('%Y年%m月%d日')} 课程考勤记录"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_alignment
    
    # 写入表头
    headers = ['课程名称', '上课时间', '上课地点', '应到人数', '实到人数', '到勤率']
    ws.append([''])  # 空行
    ws.append(headers)
    
    header_row = ws.max_row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # 写入课程数据
    total_students = Student.query.count()
    
    for course in courses:
        # 获取该课程的考勤统计
        present_count = Attendance.query.filter_by(
            course_id=course.course_id,
            attendance_date=export_date,
            attendance_type='到课'
        ).count()
        
        attendance_rate = round((present_count / total_students * 100), 1) if total_students > 0 else 0
        
        # 提取上课时间（从course_time中提取）
        time_parts = course.course_time.split()
        time_str = time_parts[1] if len(time_parts) > 1 else ''
        
        row_data = [
            course.course_name,
            time_str,
            course.course_place,
            total_students,
            present_count,
            f"{attendance_rate}%"
        ]
        ws.append(row_data)
        
        # 设置边框和对齐
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.border = border
            cell.alignment = center_alignment
    
    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{export_date.strftime('%Y%m%d')}_考勤记录.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@attendance_bp.route('/export_week')
def export_week():
    """导出当周的考勤记录"""
    # 获取日期参数
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    try:
        current_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('无效的日期格式', 'error')
        return redirect(url_for('attendance.index'))
    
    # 计算周一和周日
    week_start = current_date - timedelta(days=current_date.weekday())
    week_end = week_start + timedelta(days=6)
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "周考勤统计"
    
    # 定义样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入标题
    ws.merge_cells('A1:G1')
    ws['A1'] = f"{week_start.strftime('%Y年%m月%d日')} - {week_end.strftime('%Y年%m月%d日')} 课程考勤统计"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_alignment
    
    # 遍历一周的每一天
    current_row = 3
    total_students = Student.query.count()
    
    for day_offset in range(7):
        current_day = week_start + timedelta(days=day_offset)
        
        # 获取当天的课程
        # 根据星期几查询课程
        weekdays = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
        current_weekday = weekdays[current_day.weekday()]
        all_courses = Course.query.filter(
            Course.course_time.like(f'{current_weekday}%')
        ).order_by(Course.course_time).all()
        
        # 根据周次过滤课程
        day_week = get_week_number(current_day)
        courses = []
        if day_week:
            for course in all_courses:
                if course.is_teaching_week(day_week):
                    courses.append(course)
        else:
            courses = all_courses
        
        if not courses:
            continue
        
        # 写入日期标题
        ws.merge_cells(f'A{current_row}:G{current_row}')
        date_cell = ws.cell(row=current_row, column=1)
        date_cell.value = f"{current_day.strftime('%Y年%m月%d日')} {current_weekday}"
        date_cell.font = Font(bold=True, size=12)
        date_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        date_cell.alignment = center_alignment
        current_row += 1
        
        # 写入表头
        headers = ['课程名称', '上课时间', '上课地点', '应到', '实到', '请假', '旷课']
        ws.append(headers)
        
        header_row = current_row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        current_row += 1
        
        # 写入课程数据
        for course in courses:
            # 获取该课程的考勤统计
            present = Attendance.query.filter_by(
                course_id=course.course_id,
                attendance_date=current_day,
                attendance_type='到课'
            ).count()
            
            leave = Attendance.query.filter_by(
                course_id=course.course_id,
                attendance_date=current_day,
                attendance_type='请假'
            ).count()
            
            absent = Attendance.query.filter_by(
                course_id=course.course_id,
                attendance_date=current_day,
                attendance_type='旷课'
            ).count()
            
            # 提取上课时间
            time_parts = course.course_time.split()
            time_str = time_parts[1] if len(time_parts) > 1 else ''
            
            row_data = [
                course.course_name,
                time_str,
                course.course_place,
                total_students,
                present,
                leave,
                absent
            ]
            ws.append(row_data)
            
            # 设置边框和对齐
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.border = border
                cell.alignment = center_alignment
            
            current_row += 1
        
        # 空行
        current_row += 1
    
    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{week_start.strftime('%Y%m%d')}-{week_end.strftime('%Y%m%d')}_周考勤统计.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@attendance_bp.route('/history')
def history():
    """历史考勤查询"""
    # 获取查询参数
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')
    course_id = request.args.get('course_id', '')
    
    # 构建查询
    query = db.session.query(
        Attendance.attendance_date,
        Course.course_name,
        Course.course_id,
        func.count(Attendance.attendance_id).label('total'),
        func.sum(case((Attendance.attendance_type == '到课', 1), else_=0)).label('present'),
        func.sum(case((Attendance.attendance_type == '请假', 1), else_=0)).label('leave'),
        func.sum(case((Attendance.attendance_type == '旷课', 1), else_=0)).label('absent')
    ).join(Course).group_by(
        Attendance.attendance_date,
        Course.course_name,
        Course.course_id
    ).order_by(Attendance.attendance_date.desc())
    
    # 应用过滤条件
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            query = query.filter(Attendance.attendance_date >= start_date)
        except ValueError:
            pass
    
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            query = query.filter(Attendance.attendance_date <= end_date)
        except ValueError:
            pass
    
    if course_id:
        query = query.filter(Course.course_id == course_id)
    
    # 分页
    page = request.args.get('page', 1, type=int)
    per_page = 15
    
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    records = pagination.items
    
    # 获取所有课程用于筛选
    courses = Course.query.order_by(Course.course_name).all()
    
    # 计算到勤率并转换为字典列表
    total_students = Student.query.count()
    processed_records = []
    for record in records:
        record_dict = {
            'attendance_date': record.attendance_date,
            'course_name': record.course_name,
            'course_id': record.course_id,
            'total': record.total,
            'present': record.present,
            'leave': record.leave,
            'absent': record.absent,
            'rate': round((record.present / total_students * 100), 1) if total_students > 0 else 0
        }
        processed_records.append(record_dict)
    
    return render_template(
        'attendance/history.html',
        records=processed_records,
        pagination=pagination,
        courses=courses,
        start_date=start_date_str,
        end_date=end_date_str,
        selected_course_id=course_id,
        total_students=total_students
    )

