"""
请假管理路由模块
"""
from flask import Blueprint, render_template, request, flash, redirect, url_for, send_file
from flask_login import login_required
from datetime import datetime, date, timedelta
from sqlalchemy import func, and_, or_

from app import db
from app.models.leave_record import LeaveRecord
from app.models.student import Student

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

leave_bp = Blueprint('leave', __name__)

@leave_bp.route('/')
def index():
    """请假记录列表"""
    # 获取查询参数
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    leave_type = request.args.get('leave_type', '', type=str)
    start_date = request.args.get('start_date', '', type=str)
    end_date = request.args.get('end_date', '', type=str)
    
    # 构建查询
    query = LeaveRecord.query.join(Student)
    
    # 搜索条件（学号或姓名）
    if search:
        search_pattern = f'%{search}%'
        query = query.filter(
            or_(
                Student.student_id.like(search_pattern),
                Student.student_name.like(search_pattern)
            )
        )
    
    # 请假类型筛选
    if leave_type:
        query = query.filter(LeaveRecord.leave_type == leave_type)
    
    # 日期范围筛选
    if start_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(LeaveRecord.leave_start_date >= start)
        except ValueError:
            pass
    
    if end_date:
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(LeaveRecord.leave_end_date <= end)
        except ValueError:
            pass
    
    # 按创建时间倒序排列
    query = query.order_by(LeaveRecord.create_time.desc())
    
    # 分页
    per_page = 15
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    records = pagination.items
    
    # 获取所有请假类型（用于筛选下拉框）
    leave_types = db.session.query(LeaveRecord.leave_type).distinct().all()
    leave_types = [t[0] for t in leave_types]
    
    return render_template(
        'leave/index.html',
        records=records,
        pagination=pagination,
        search=search,
        leave_type=leave_type,
        start_date=start_date,
        end_date=end_date,
        leave_types=leave_types
    )


@leave_bp.route('/add', methods=['GET', 'POST'])
def add():
    """添加请假记录"""
    if request.method == 'POST':
        student_id = request.form.get('student_id', '').strip()
        leave_type = request.form.get('leave_type', '').strip()
        leave_start_date = request.form.get('leave_start_date', '').strip()
        leave_end_date = request.form.get('leave_end_date', '').strip()
        leave_reason = request.form.get('leave_reason', '').strip()
        
        # 验证必填字段
        if not all([student_id, leave_type, leave_start_date, leave_end_date, leave_reason]):
            flash('请填写所有必填字段', 'error')
            return redirect(url_for('leave.add'))
        
        # 验证学生是否存在
        student = Student.query.get(student_id)
        if not student:
            flash('学生不存在', 'error')
            return redirect(url_for('leave.add'))
        
        # 验证日期
        try:
            start_date = datetime.strptime(leave_start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(leave_end_date, '%Y-%m-%d').date()
            
            if end_date < start_date:
                flash('结束日期不能早于开始日期', 'error')
                return redirect(url_for('leave.add'))
            
            # 计算请假天数
            leave_days = (end_date - start_date).days + 1
            
        except ValueError:
            flash('日期格式错误', 'error')
            return redirect(url_for('leave.add'))
        
        # 创建请假记录
        leave_record = LeaveRecord(
            student_id=student_id,
            leave_type=leave_type,
            leave_start_date=start_date,
            leave_end_date=end_date,
            leave_days=leave_days,
            leave_reason=leave_reason
        )
        
        try:
            db.session.add(leave_record)
            db.session.commit()
            flash('请假记录添加成功', 'success')
            return redirect(url_for('leave.index'))
        except Exception as e:
            db.session.rollback()
            flash(f'添加失败：{str(e)}', 'error')
            return redirect(url_for('leave.add'))
    
    # GET请求：显示添加表单
    students = Student.query.order_by(Student.student_id).all()
    return render_template('leave/form.html', students=students, record=None)


@leave_bp.route('/edit/<int:leave_id>', methods=['GET', 'POST'])
def edit(leave_id):
    """编辑请假记录"""
    record = LeaveRecord.query.get_or_404(leave_id)
    
    if request.method == 'POST':
        student_id = request.form.get('student_id', '').strip()
        leave_type = request.form.get('leave_type', '').strip()
        leave_start_date = request.form.get('leave_start_date', '').strip()
        leave_end_date = request.form.get('leave_end_date', '').strip()
        leave_reason = request.form.get('leave_reason', '').strip()
        
        # 验证必填字段
        if not all([student_id, leave_type, leave_start_date, leave_end_date, leave_reason]):
            flash('请填写所有必填字段', 'error')
            return redirect(url_for('leave.edit', leave_id=leave_id))
        
        # 验证学生是否存在
        student = Student.query.get(student_id)
        if not student:
            flash('学生不存在', 'error')
            return redirect(url_for('leave.edit', leave_id=leave_id))
        
        # 验证日期
        try:
            start_date = datetime.strptime(leave_start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(leave_end_date, '%Y-%m-%d').date()
            
            if end_date < start_date:
                flash('结束日期不能早于开始日期', 'error')
                return redirect(url_for('leave.edit', leave_id=leave_id))
            
            # 计算请假天数
            leave_days = (end_date - start_date).days + 1
            
        except ValueError:
            flash('日期格式错误', 'error')
            return redirect(url_for('leave.edit', leave_id=leave_id))
        
        # 更新记录
        record.student_id = student_id
        record.leave_type = leave_type
        record.leave_start_date = start_date
        record.leave_end_date = end_date
        record.leave_days = leave_days
        record.leave_reason = leave_reason
        
        try:
            db.session.commit()
            flash('请假记录更新成功', 'success')
            return redirect(url_for('leave.index'))
        except Exception as e:
            db.session.rollback()
            flash(f'更新失败：{str(e)}', 'error')
            return redirect(url_for('leave.edit', leave_id=leave_id))
    
    # GET请求：显示编辑表单
    students = Student.query.order_by(Student.student_id).all()
    return render_template('leave/form.html', students=students, record=record)


@leave_bp.route('/delete/<int:leave_id>', methods=['POST'])
def delete(leave_id):
    """删除请假记录"""
    record = LeaveRecord.query.get_or_404(leave_id)
    
    try:
        db.session.delete(record)
        db.session.commit()
        flash('请假记录删除成功', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败：{str(e)}', 'error')
    
    return redirect(url_for('leave.index'))


@leave_bp.route('/statistics')
def statistics():
    """请假统计"""
    # 获取筛选参数
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')
    
    # 默认统计本月数据
    if not start_date_str or not end_date_str:
        today = date.today()
        start_date = date(today.year, today.month, 1)
        # 下月第一天的前一天
        if today.month == 12:
            end_date = date(today.year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(today.year, today.month + 1, 1) - timedelta(days=1)
    else:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            flash('日期格式错误', 'error')
            return redirect(url_for('leave.statistics'))
    
    # 构建基础查询
    base_query = LeaveRecord.query.filter(
        LeaveRecord.leave_start_date <= end_date,
        LeaveRecord.leave_end_date >= start_date
    )
    
    # 总体统计
    total_count = base_query.count()
    total_days = db.session.query(func.sum(LeaveRecord.leave_days)).filter(
        LeaveRecord.leave_start_date <= end_date,
        LeaveRecord.leave_end_date >= start_date
    ).scalar() or 0
    
    # 按请假类型统计
    type_stats = db.session.query(
        LeaveRecord.leave_type,
        func.count(LeaveRecord.leave_id).label('count'),
        func.sum(LeaveRecord.leave_days).label('days')
    ).filter(
        LeaveRecord.leave_start_date <= end_date,
        LeaveRecord.leave_end_date >= start_date
    ).group_by(LeaveRecord.leave_type).all()
    
    # 按学生统计（请假次数最多的前10名）
    student_stats = db.session.query(
        Student.student_id,
        Student.student_name,
        func.count(LeaveRecord.leave_id).label('count'),
        func.sum(LeaveRecord.leave_days).label('days')
    ).join(LeaveRecord).filter(
        LeaveRecord.leave_start_date <= end_date,
        LeaveRecord.leave_end_date >= start_date
    ).group_by(
        Student.student_id,
        Student.student_name
    ).order_by(func.count(LeaveRecord.leave_id).desc()).limit(10).all()
    
    # 最近的请假记录
    recent_records = base_query.order_by(LeaveRecord.create_time.desc()).limit(10).all()
    
    return render_template(
        'leave/statistics.html',
        start_date=start_date,
        end_date=end_date,
        total_count=total_count,
        total_days=total_days,
        type_stats=type_stats,
        student_stats=student_stats,
        recent_records=recent_records
    )


@leave_bp.route('/export')
def export():
    """导出请假记录"""
    # 获取筛选参数
    search = request.args.get('search', '', type=str)
    leave_type = request.args.get('leave_type', '', type=str)
    start_date_str = request.args.get('start_date', '', type=str)
    end_date_str = request.args.get('end_date', '', type=str)
    
    # 构建查询
    query = LeaveRecord.query.join(Student)
    
    if search:
        search_pattern = f'%{search}%'
        query = query.filter(
            or_(
                Student.student_id.like(search_pattern),
                Student.student_name.like(search_pattern)
            )
        )
    
    if leave_type:
        query = query.filter(LeaveRecord.leave_type == leave_type)
    
    if start_date_str:
        try:
            start = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            query = query.filter(LeaveRecord.leave_start_date >= start)
        except ValueError:
            pass
    
    if end_date_str:
        try:
            end = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            query = query.filter(LeaveRecord.leave_end_date <= end)
        except ValueError:
            pass
    
    records = query.order_by(LeaveRecord.create_time.desc()).all()
    
    if not records:
        flash('没有符合条件的请假记录', 'warning')
        return redirect(url_for('leave.index'))
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "请假记录"
    
    # 定义样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入标题
    ws.merge_cells('A1:H1')
    ws['A1'] = "请假记录导出"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_alignment
    
    # 写入表头
    headers = ['学号', '姓名', '请假类型', '开始日期', '结束日期', '请假天数', '请假原因', '提交时间']
    ws.append([''])  # 空行
    ws.append(headers)
    
    header_row = ws.max_row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # 写入数据
    for record in records:
        row_data = [
            record.student.student_id,
            record.student.student_name,
            record.leave_type,
            record.leave_start_date.strftime('%Y-%m-%d'),
            record.leave_end_date.strftime('%Y-%m-%d'),
            record.leave_days,
            record.leave_reason,
            record.create_time.strftime('%Y-%m-%d %H:%M')
        ]
        ws.append(row_data)
        
        # 设置边框和对齐
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.border = border
            if col_num <= 6:  # 前6列居中
                cell.alignment = center_alignment
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 18
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"请假记录_{timestamp}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

