from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required
from app import db
from app.models import Course, Attendance, Student
from werkzeug.utils import secure_filename
from icalendar import Calendar
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import re

# 创建课程管理蓝图
course_bp = Blueprint('course', __name__)

# 允许的文件扩展名
ALLOWED_EXTENSIONS = {'ics'}

def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _format_week_range(weeks):
    """
    将周次列表转换为范围字符串
    
    Args:
        weeks: 周次列表（已排序）
        
    Returns:
        str: 周次范围字符串（如：1-8,13,15,17 或 9-16）
    """
    if not weeks:
        return None
    
    ranges = []
    start = weeks[0]
    end = weeks[0]
    
    for i in range(1, len(weeks)):
        if weeks[i] == end + 1:
            end = weeks[i]
        else:
            if start == end:
                ranges.append(str(start))
            else:
                ranges.append(f"{start}-{end}")
            start = weeks[i]
            end = weeks[i]
    
    # 处理最后一个范围
    if start == end:
        ranges.append(str(start))
    else:
        ranges.append(f"{start}-{end}")
    
    return ','.join(ranges)


@course_bp.route('/')
@login_required
def index():
    """课程列表页面"""
    # 获取查询参数
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    semester = request.args.get('semester', '', type=str)
    
    # 构建查询
    query = Course.query
    
    # 搜索条件
    if search:
        query = query.filter(
            (Course.course_id.contains(search)) |
            (Course.course_name.contains(search)) |
            (Course.teacher_name.contains(search))
        )
    
    # 学期筛选
    if semester:
        query = query.filter(Course.semester == semester)
    
    # 分页查询
    pagination = query.order_by(Course.course_id).paginate(
        page=page,
        per_page=10,
        error_out=False
    )
    
    courses = pagination.items
    
    # 为每个课程计算到勤率
    for course in courses:
        total_attendances = Attendance.query.filter_by(course_id=course.course_id).count()
        if total_attendances > 0:
            normal_attendances = Attendance.query.filter_by(
                course_id=course.course_id,
                attendance_type='正常'
            ).count()
            course.attendance_rate = round((normal_attendances / total_attendances) * 100, 1)
        else:
            course.attendance_rate = 0
        course.total_records = total_attendances
    
    # 获取所有学期选项
    semesters = db.session.query(Course.semester)\
        .distinct()\
        .filter(Course.semester.isnot(None))\
        .order_by(Course.semester.desc())\
        .all()
    semesters = [s[0] for s in semesters]
    
    return render_template(
        'course/index.html',
        courses=courses,
        pagination=pagination,
        search=search,
        semester=semester,
        semesters=semesters
    )


@course_bp.route('/import', methods=['GET', 'POST'])
@login_required
def import_ics():
    """从ICS文件导入课程"""
    if request.method == 'POST':
        # 检查是否有文件上传
        if 'file' not in request.files:
            flash('没有选择文件！', 'error')
            return redirect(url_for('course.import_ics'))
        
        file = request.files['file']
        
        if file.filename == '':
            flash('没有选择文件！', 'error')
            return redirect(url_for('course.import_ics'))
        
        if not allowed_file(file.filename):
            flash('仅支持 .ics 格式的日历文件！', 'error')
            return redirect(url_for('course.import_ics'))
        
        try:
            # 解析ICS文件
            cal = Calendar.from_ical(file.read())
            
            # 用于存储课程信息的字典（按课程代码分组）
            courses_dict = {}
            
            # 遍历所有事件
            for component in cal.walk():
                if component.name == "VEVENT":
                    summary = str(component.get('summary'))
                    description = str(component.get('description', ''))
                    location = str(component.get('location', ''))
                    dtstart = component.get('dtstart').dt
                    dtend = component.get('dtend').dt
                    
                    # 从描述中提取课程代码和教师
                    course_code_match = re.search(r'课程代码:\s*(\w+)', description)
                    teacher_match = re.search(r'教师:\s*([^\s]+)', description)
                    week_match = re.search(r'周次:\s*(\d+)', description)
                    
                    if course_code_match:
                        course_code = course_code_match.group(1)
                        teacher_name = teacher_match.group(1) if teacher_match else '未知'
                        week_number = int(week_match.group(1)) if week_match else None
                        
                        # 提取课程名称（移除必修/选修标识）
                        course_name = re.sub(r'[（(](必修课|选修课|限选课)[）)]', '', summary).strip()
                        
                        # 提取上课时间信息
                        if isinstance(dtstart, datetime):
                            weekday_names = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
                            weekday = weekday_names[dtstart.weekday()]
                            start_time = dtstart.strftime('%H:%M')
                            end_time = dtend.strftime('%H:%M')
                            course_time = f"{weekday} {start_time}-{end_time}"
                        else:
                            course_time = "待定"
                        
                        # 为每个"课程代码+上课时间"创建唯一标识
                        # 这样一周多次上课的课程会创建多条记录
                        # 格式：课程代码-星期-时间（如：CS101-1-0800）
                        weekday_num = dtstart.weekday() + 1 if isinstance(dtstart, datetime) else 0
                        time_code = start_time.replace(':', '') if isinstance(dtstart, datetime) else '0000'
                        unique_key = f"{course_code}-{weekday_num}-{time_code}"
                        
                        # 如果该课程+时间组合不存在，则添加
                        if unique_key not in courses_dict:
                            courses_dict[unique_key] = {
                                'course_id': unique_key,  # 使用唯一标识作为ID
                                'course_name': course_name,
                                'teacher_name': teacher_name,
                                'course_time': course_time,
                                'course_place': location,
                                'semester': '2025-2026学年第一学期',  # 从文件名或描述中提取
                                'weeks': set()  # 用于收集所有周次
                            }
                        
                        # 收集周次信息
                        if week_number:
                            courses_dict[unique_key]['weeks'].add(week_number)
            
            # 处理周次信息，转换为week_range字符串
            for course_data in courses_dict.values():
                weeks = sorted(course_data.pop('weeks', set()))
                if weeks:
                    # 将周次列表转换为范围字符串
                    week_range = _format_week_range(weeks)
                    course_data['week_range'] = week_range
                else:
                    course_data['week_range'] = None
            
            # 导入数据库
            success_count = 0
            update_count = 0
            
            for course_data in courses_dict.values():
                existing_course = Course.query.filter_by(course_id=course_data['course_id']).first()
                
                if existing_course:
                    # 更新现有课程
                    existing_course.course_name = course_data['course_name']
                    existing_course.teacher_name = course_data['teacher_name']
                    existing_course.course_time = course_data['course_time']
                    existing_course.course_place = course_data['course_place']
                    existing_course.semester = course_data['semester']
                    existing_course.week_range = course_data['week_range']
                    update_count += 1
                else:
                    # 创建新课程
                    course = Course(**course_data)
                    db.session.add(course)
                    success_count += 1
            
            db.session.commit()
            
            if update_count > 0:
                flash(f'导入成功！新增 {success_count} 门课程，更新 {update_count} 门课程', 'success')
            else:
                flash(f'导入成功！共导入 {success_count} 门课程', 'success')
            
            return redirect(url_for('course.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'导入失败：{str(e)}', 'error')
            return redirect(url_for('course.import_ics'))
    
    return render_template('course/import.html')


@course_bp.route('/detail/<course_id>')
@login_required
def detail(course_id):
    """课程详情和考勤统计"""
    course = Course.query.get_or_404(course_id)
    
    # 获取日期筛选参数
    filter_date = request.args.get('date', '', type=str)
    
    # 基础查询
    attendance_query = Attendance.query.filter_by(course_id=course_id)
    
    # 日期筛选
    if filter_date:
        try:
            target_date = datetime.strptime(filter_date, '%Y-%m-%d').date()
            attendance_query = attendance_query.filter(Attendance.attendance_date == target_date)
        except ValueError:
            pass
    
    # 获取考勤记录
    attendances = attendance_query.order_by(Attendance.attendance_date.desc(), Attendance.student_id).all()
    
    # 统计信息
    total_records = Attendance.query.filter_by(course_id=course_id).count()
    
    stats = {
        'total': total_records,
        'normal': Attendance.query.filter_by(course_id=course_id, attendance_type='正常').count(),
        'late': Attendance.query.filter_by(course_id=course_id, attendance_type='迟到').count(),
        'early_leave': Attendance.query.filter_by(course_id=course_id, attendance_type='早退').count(),
        'absent': Attendance.query.filter_by(course_id=course_id, attendance_type='缺席').count(),
        'leave': Attendance.query.filter_by(course_id=course_id, attendance_type='请假').count(),
    }
    
    # 计算出勤率
    if total_records > 0:
        stats['attendance_rate'] = round((stats['normal'] / total_records) * 100, 1)
    else:
        stats['attendance_rate'] = 0
    
    # 获取所有考勤日期（用于日期筛选）
    attendance_dates = db.session.query(Attendance.attendance_date)\
        .filter_by(course_id=course_id)\
        .distinct()\
        .order_by(Attendance.attendance_date.desc())\
        .all()
    attendance_dates = [d[0] for d in attendance_dates]
    
    return render_template(
        'course/detail.html',
        course=course,
        attendances=attendances,
        stats=stats,
        filter_date=filter_date,
        attendance_dates=attendance_dates
    )


@course_bp.route('/export_attendance')
@login_required
def export_attendance():
    """导出课程考勤数据"""
    course_id = request.args.get('course_id', '', type=str)
    filter_date = request.args.get('date', '', type=str)
    
    if not course_id:
        flash('请指定课程！', 'error')
        return redirect(url_for('course.index'))
    
    course = Course.query.get_or_404(course_id)
    
    # 构建查询
    query = Attendance.query.filter_by(course_id=course_id)
    
    # 日期筛选
    filename_suffix = ''
    if filter_date:
        try:
            target_date = datetime.strptime(filter_date, '%Y-%m-%d').date()
            query = query.filter(Attendance.attendance_date == target_date)
            filename_suffix = f'_{filter_date}'
        except ValueError:
            pass
    
    attendances = query.order_by(Attendance.attendance_date, Attendance.student_id).all()
    
    if not attendances:
        flash('没有考勤数据可导出！', 'warning')
        return redirect(url_for('course.detail', course_id=course_id))
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "考勤记录"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入课程信息
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = f"《{course.course_name}》考勤记录"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells('A2:G2')
    info_cell = ws['A2']
    info_cell.value = f"教师：{course.teacher_name}  |  上课时间：{course.course_time}  |  地点：{course.course_place}"
    info_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入表头
    headers = ["学号", "姓名", "考勤日期", "考勤状态", "迟到分钟", "备注", "记录时间"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15  # 学号
    ws.column_dimensions['B'].width = 12  # 姓名
    ws.column_dimensions['C'].width = 15  # 考勤日期
    ws.column_dimensions['D'].width = 12  # 考勤状态
    ws.column_dimensions['E'].width = 12  # 迟到分钟
    ws.column_dimensions['F'].width = 30  # 备注
    ws.column_dimensions['G'].width = 20  # 记录时间
    
    # 写入数据
    data_alignment = Alignment(horizontal="left", vertical="center")
    for row_num, attendance in enumerate(attendances, 5):
        ws.cell(row=row_num, column=1).value = attendance.student_id
        ws.cell(row=row_num, column=2).value = attendance.student.student_name
        ws.cell(row=row_num, column=3).value = attendance.attendance_date.strftime('%Y-%m-%d')
        
        # 考勤状态单元格添加颜色
        status_cell = ws.cell(row=row_num, column=4)
        status_cell.value = attendance.attendance_type
        if attendance.attendance_type == '正常':
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif attendance.attendance_type == '迟到':
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif attendance.attendance_type == '缺席':
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        ws.cell(row=row_num, column=5).value = attendance.late_minutes if attendance.late_minutes else 0
        ws.cell(row=row_num, column=6).value = attendance.attendance_note or '-'
        ws.cell(row=row_num, column=7).value = attendance.create_time.strftime('%Y-%m-%d %H:%M:%S') if attendance.create_time else '-'
        
        # 设置数据行对齐
        for col_num in range(1, 8):
            ws.cell(row=row_num, column=col_num).alignment = data_alignment
    
    # 生成文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'{course.course_name}_考勤记录{filename_suffix}_{timestamp}.xlsx'
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 返回文件
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@course_bp.route('/delete/<course_id>', methods=['POST'])
@login_required
def delete(course_id):
    """删除课程"""
    course = Course.query.get_or_404(course_id)
    course_name = course.course_name
    
    try:
        db.session.delete(course)
        db.session.commit()
        flash(f'课程《{course_name}》删除成功！', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败：{str(e)}', 'error')
    
    return redirect(url_for('course.index'))


@course_bp.route('/today')
@login_required
def today_attendance():
    """查看今日所有课程考勤"""
    today = date.today()
    
    # 获取今日所有考勤记录
    attendances = Attendance.query.filter_by(attendance_date=today)\
        .order_by(Attendance.course_id, Attendance.student_id)\
        .all()
    
    # 按课程分组统计
    course_stats = {}
    for attendance in attendances:
        course_id = attendance.course_id
        if course_id not in course_stats:
            course_stats[course_id] = {
                'course': attendance.course,
                'total': 0,
                'present': 0,
                'leave': 0,
                'absent': 0,
                'records': []
            }
        
        course_stats[course_id]['total'] += 1
        if attendance.attendance_type == '到课':
            course_stats[course_id]['present'] += 1
        elif attendance.attendance_type == '请假':
            course_stats[course_id]['leave'] += 1
        elif attendance.attendance_type == '旷课':
            course_stats[course_id]['absent'] += 1
        
        course_stats[course_id]['records'].append(attendance)
    
    # 计算出勤率
    for stats in course_stats.values():
        if stats['total'] > 0:
            stats['attendance_rate'] = round((stats['present'] / stats['total']) * 100, 1)
        else:
            stats['attendance_rate'] = 0
    
    return render_template(
        'course/today.html',
        today=today,
        course_stats=course_stats
    )

